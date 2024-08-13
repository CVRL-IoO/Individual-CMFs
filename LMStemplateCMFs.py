# -*- coding: utf-8 -*-
"""
Created on Tue Aug 09 13:23:40 2016

@author: Caesar

Related files:
*.ui  Qt5 designer files defining GUI components and layouts

/CMFs_out directory required for data output
"""

import time, sys, os, ctypes

import numpy as np

import openpyxl #For writing Excel files

import CMFcalc, CMFtemplates #CMF modules

from CMFplot import CMFPlot #CMF plotting class

from PyQt5 import QtWidgets, QtCore, uic #pyqt stuff

 
QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True) #enable highdpi scaling
QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps, True) #use highdpi icons



class individualtemplatesCMFs(QtWidgets.QMainWindow):
    
    def __init__(self):
        
        QtWidgets.QMainWindow.__init__(self)
        
        #Main CSF window
        #For qt window scaling, environment variable: QT_AUTO_SCREEN_SCALE_FACTOR=0 () in winpython.ini
        
        self.scaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0) #Windows optimize GUI for different Windows scale factors
        self.mainCMFgen = uic.loadUi('inputCMFgenerate.ui')  #GUI screen (defaults)
        
        #Initialise window position and size depending on screen size and scaling
        if screenheight/self.scaleFactor*100.0>1080:
            winheight=1000
            winwidth=1000
        else:
            winheight=round(screenheight*0.9)
            winwidth=round(screenheight*0.9)
            
        self.mainCMFgen.resize(winwidth,winheight)
        self.mainCMFgen.move(QtWidgets.QDesktopWidget().availableGeometry().center() - self.mainCMFgen.frameGeometry().center())
        self.mainCMFgen.show()
        
        self.nicedateStr = time.strftime("%H:%M %B %d, %Y", time.localtime())
        #Standard quantal lmax values for individual log-lin 5x2 absorbance templates (closest 0.01 nm values)
       
        self.Lserlmax_template=554.86   #Peak at 0.01 step
        self.Lalalmax_template=550.85
        self.Lmeanlmax_template=553.09
        self.Mlmax_template=531.19
        self.Slmax_template=419.00
        self.Lser_Mlmax_diff=self.Lserlmax_template-self.Mlmax_template #Currently 23.67


        #StockmanSharpe/CIE assumed mean 2 deg photopigment optical densities; new assumption
        self.Lod_norm_2 = 0.50
        self.Mod_norm_2 = 0.50
        self.Sod_norm_2 = 0.40

        #Stiles & Burch primaries
        self.SB_Rnm = 645.15
        self.SB_Gnm = 526.32
        self.SB_Bnm = 444.44

        #StockmanSharpe assumed mean 10-deg photopigment optical densities. For reference (not used).
        self.Lod_norm_10 = 0.38
        self.Mod_norm_10 = 0.38
        self.Sod_norm_10 = 0.30

        #StockmanSharpe/CIE assumed mean macular densities for 2 and 10 deg at 460 nm, and lens densities at 400 nm
        self.mac_norm_2_460 = 0.350 #0.35 at 460 nm
        self.mac_norm_10_460 = 0.095   #For reference (not used).
        self.lens_norm_400 = 1.7649  #1.7649 Stockman & Sharpe

        #Codon shifts for M->L Edit here for changes
        self.M_L_116baseshift = 0
        self.M_L_180baseshift = 3
        self.M_L_230baseshift = 3
        self.M_L_233baseshift = 0
        self.M_L_277baseshift = 7
        self.M_L_285baseshift = 14
        self.M_L_309baseshift = 0
        self.M_L_allbaseshifts = (self.M_L_116baseshift + self.M_L_180baseshift + self.M_L_230baseshift 
                                    + self.M_L_233baseshift + self.M_L_277baseshift + self.M_L_285baseshift
                                    + self.M_L_309baseshift)
        #Total shift is 28 nm but scaled up by 23.67/28 to equal 23.67 nm
        
        self.M_L_scale= self.Lser_Mlmax_diff/self.M_L_allbaseshifts

        self.M_L_116shift = self.M_L_116baseshift * self.M_L_scale
        self.M_L_180shift = self.M_L_180baseshift * self.M_L_scale
        self.M_L_230shift = self.M_L_230baseshift * self.M_L_scale
        self.M_L_233shift = self.M_L_233baseshift * self.M_L_scale
        self.M_L_277shift = self.M_L_277baseshift * self.M_L_scale
        self.M_L_285shift = self.M_L_285baseshift * self.M_L_scale
        self.M_L_309shift = self.M_L_309baseshift * self.M_L_scale
        
        
        #Codon shifts for L->M Edit here for changes
        self.L_M_116baseshift = -3
        self.L_M_180baseshift = -4
        self.L_M_230baseshift = -3
        self.L_M_233baseshift =  0
        self.L_M_277baseshift = -7
        self.L_M_285baseshift = -14
        self.L_M_309baseshift =  0
        self.L_M_allbaseshifts = (self.L_M_116baseshift + self.L_M_180baseshift + self.L_M_230baseshift 
                                    + self.L_M_233baseshift + self.L_M_277baseshift + self.L_M_285baseshift
                                    + self.L_M_309baseshift)
       
        self.L_M_scale= -self.Lser_Mlmax_diff/self.L_M_allbaseshifts
       #Total shift is 23.67 nm -- too much. Scaled by 23.67/31 so total shift is 23.67 

        self.L_M_116shift = self.L_M_116baseshift * self.L_M_scale
        self.L_M_180shift = self.L_M_180baseshift * self.L_M_scale
        self.L_M_230shift = self.L_M_230baseshift * self.L_M_scale
        self.L_M_233shift = self.L_M_233baseshift * self.L_M_scale
        self.L_M_277shift = self.L_M_277baseshift * self.L_M_scale
        self.L_M_285shift = self.L_M_285baseshift * self.L_M_scale
        self.L_M_309shift = self.L_M_309baseshift * self.L_M_scale
       
        #Flags for using L cone templates for M or M for L, initially false! U        
        self.MisL = False
        self.LisM = False       
          
        self.useCodons = False #Flag set when using codons to calculate lmax
        
        #L and M-cone shifts, set to 0 to start
        self.M_shift = 0
        self.L_shift = 0
        self.ML_shift = 0  #For codon screen calculations
        self.LM_shift = 0  #For codon screen calculations


        #Blank some messages and warnings and buttons for later use
        self.mainCMFgen.Commontxt.hide() #Hide Common template notice leaving Individual template notice.
        self.mainCMFgen.useMforL.setVisible(False)  #Notification that M-cone template is being used for L (hidden)
        self.mainCMFgen.useLforM.setVisible(False)  #Notification that L-cone template is being used for M (hidden)

        self.mainCMFgen.notuseCodons.hide() #Button to revert to initial default of using direct input for L and M shifts
 
        self.LM_codons = uic.loadUi('LM_codons.ui') #Second codon input screen
        if screenheight/self.scaleFactor*100.0>580:
            self.LM_codons.setGeometry(200,50,620,580)
        else:
            self.LM_codons.setGeometry(200,50,round(620/self.scaleFactor*100.0),round(580/self.scaleFactor*100.0))
        
            
        allmainFrames = self.mainCMFgen.findChildren(QtWidgets.QFrame)

        for thisFrame in allmainFrames:
            self.scaleThing(thisFrame,self.mainCMFgen.geometry().height()/1000)
            
        allmainButtons = self.mainCMFgen.findChildren(QtWidgets.QPushButton)
        for thisButton in allmainButtons:
            self.scaleThing(thisButton,self.mainCMFgen.geometry().height()/1000)
            
        # allmainLabels = self.mainCMFgen.findChildren(QtWidgets.QLabel)#all QLabel objects get moved/resized by QFrame above
        # for thisLabel in allmainLabels:
        #     self.scaleThing(thisLabel,self.mainCMFgen.geometry().height()/1000)
            
        allmainLEdits = self.mainCMFgen.findChildren(QtWidgets.QLineEdit)
        for thisEdit in allmainLEdits:
            self.scaleThing(thisEdit,self.mainCMFgen.geometry().height()/1000)
            
        allmainChecks = self.mainCMFgen.findChildren(QtWidgets.QCheckBox)
        for thisCheck in allmainChecks:
            self.scaleThing(thisCheck,self.mainCMFgen.geometry().height()/1000)
            
        allmainRadios = self.mainCMFgen.findChildren(QtWidgets.QRadioButton)
        for thisRadio in allmainRadios:
            self.scaleThing(thisRadio,self.mainCMFgen.geometry().height()/1000)
            
        allmainGroups = self.mainCMFgen.findChildren(QtWidgets.QGroupBox)
        for thisGroup in allmainGroups:
            self.scaleThing(thisGroup,self.mainCMFgen.geometry().height()/1000)
        
        
        allCodFrames = self.LM_codons.findChildren(QtWidgets.QFrame)
        for thisFrame in allCodFrames:
            self.scaleThing(thisFrame,self.LM_codons.geometry().height()/580)
            
        allCodButtons = self.LM_codons.findChildren(QtWidgets.QPushButton)
        for thisButton in allCodButtons:
            self.scaleThing(thisButton,self.LM_codons.geometry().height()/580)
            
        # allCodLabels = self.LM_codons.findChildren(QtWidgets.QLabel)#all QLabel objects get moved/resized by QFrame above
        # for thisLabel in allCodLabels:
        #     self.scaleThing(thisLabel,self.LM_codons.geometry().height()/580)
            
        allCodLEdits = self.LM_codons.findChildren(QtWidgets.QLineEdit)
        for thisEdit in allCodLEdits:
            self.scaleThing(thisEdit,self.LM_codons.geometry().height()/580)
            
        allCodChecks = self.LM_codons.findChildren(QtWidgets.QCheckBox)
        for thisCheck in allCodChecks:
            self.scaleThing(thisCheck,self.LM_codons.geometry().height()/580)
            
        allCodRadios = self.LM_codons.findChildren(QtWidgets.QRadioButton)
        for thisRadio in allCodRadios:
            self.scaleThing(thisRadio,self.LM_codons.geometry().height()/580)
            
        allCodGroups = self.LM_codons.findChildren(QtWidgets.QGroupBox)
        for thisGroup in allCodGroups:
            self.scaleThing(thisGroup,self.LM_codons.geometry().height()/580)
            
        self.connectButtons()
        
        self.connectRadioButtons()
        
        self.initCMFsInputScreen()
        
        self.mainCMFgen.CMFinfo = self.readFromGui()

 
    def scaleThing(self,thing,alpha):

        if self.scaleFactor>140:
            fontS=200/self.scaleFactor*alpha
        else:
            fontS=100/self.scaleFactor*alpha
        thisFont=thing.font()
        thisFSize=thisFont.pointSize()
        thisFont.setPointSize(int(thisFSize*fontS))
        thing.setFont(thisFont)
        thingLeft = thing.geometry().left()
        thingTop = thing.geometry().top()
        thingWidth = thing.geometry().width()
        thingHeight = thing.geometry().height()
        thing.setGeometry(round(thingLeft*alpha),round(thingTop*alpha),round(thingWidth*alpha),round(thingHeight*alpha))

        
    def connectButtons(self):
        self.mainCMFgen.GenerateData.clicked.connect(self.slotOutputCMFsClicked)
        self.mainCMFgen.GenerateMeanLMS.clicked.connect(self.slotOutputMeanLMSCMFsClicked)
        self.mainCMFgen.PlotCMFs.clicked.connect(self.slotPlotCMFsClicked)
        self.mainCMFgen.useCodons.clicked.connect(self.slotuseCodonsClicked)
        self.mainCMFgen.notuseCodons.clicked.connect(self.slotnotuseCodonsClicked)
        self.LM_codons.codonsDone.clicked.connect(self.slotcodonsDoneClicked)
        self.mainCMFgen.reset.clicked.connect(self.slotresetClicked)
        self.mainCMFgen.deg2Densities.clicked.connect(self.deg2DensitiesClicked)
        self.mainCMFgen.deg10Densities.clicked.connect(self.deg10DensitiesClicked)


    def connectRadioButtons(self):
        #Change M buttons. Pairs are exclusive so only need watch one.
        self.LM_codons.M_116_Ser.toggled.connect(self.radioM_116_Ser_changed)
        self.LM_codons.M_180_Ser.toggled.connect(self.radioM_180_Ser_changed)
        self.LM_codons.M_230_Ile.toggled.connect(self.radioM_230_Ile_changed)
        self.LM_codons.M_233_Ala.toggled.connect(self.radioM_233_Ala_changed)
        self.LM_codons.M_277_Tyr.toggled.connect(self.radioM_277_Tyr_changed)
        self.LM_codons.M_285_Thr.toggled.connect(self.radioM_285_Thr_changed)
        self.LM_codons.M_309_Tyr.toggled.connect(self.radioM_309_Tyr_changed)
        #Change L buttons
        self.LM_codons.L_116_Tyr.toggled.connect(self.radioL_116_Tyr_changed)
        self.LM_codons.L_180_Ala.toggled.connect(self.radioL_180_Ala_changed)
        self.LM_codons.L_230_Thr.toggled.connect(self.radioL_230_Thr_changed)
        self.LM_codons.L_233_Ser.toggled.connect(self.radioL_233_Ser_changed)
        self.LM_codons.L_277_Tyr.toggled.connect(self.radioL_277_Phe_changed)
        self.LM_codons.L_285_Ala.toggled.connect(self.radioL_285_Ala_changed)
        self.LM_codons.L_309_Phe.toggled.connect(self.radioL_309_Phe_changed)
 
 
    def initCMFsInputScreen(self):
     #2 deg values assumed by Stockman Sharpe loaded into GUI window
        self.mainCMFgen.Lod.setText(str(self.Lod_norm_2))
        self.mainCMFgen.Mod.setText(str(self.Mod_norm_2))
        self.mainCMFgen.Sod.setText(str(self.Sod_norm_2))
        self.mainCMFgen.mac.setText(str(self.mac_norm_2_460))
        self.mainCMFgen.lens.setText(str(self.lens_norm_400))
      #Zero shifts loaded into window. S cone spectral sensitivity assumed to be fixed.
        self.mainCMFgen.Lshift.setText(str(0.0))
        self.mainCMFgen.Mshift.setText(str(0.0))
        
        self.mainCMFgen.Rnm.setText(str(self.SB_Rnm))
        self.mainCMFgen.Gnm.setText(str(self.SB_Gnm))
        self.mainCMFgen.Bnm.setText(str(self.SB_Bnm))

      
    def readFromGui(self):
        #Read from GUI input screen and put in Dictionary called CMFinfo
        CMFinfo={}

        if self.mainCMFgen.step0_1.isChecked():     #Extract nm step-size
            step=0.1
        elif self.mainCMFgen.step1_0.isChecked():
            step=1.0
        elif self.mainCMFgen.step5_0.isChecked():
            step=5.0
        CMFinfo['step_size']=step                   #nm step size 0.1, 1 or 5 (default 1)
                
        CMFinfo['Lod'] = float(self.mainCMFgen.Lod.text())  #Optical densities
        CMFinfo['Mod'] = float(self.mainCMFgen.Mod.text())
        CMFinfo['Sod'] = float(self.mainCMFgen.Sod.text())
 
        CMFinfo['Lshift'] = float(self.mainCMFgen.Lshift.text())    #L and M shifts
        self.L_shift = CMFinfo['Lshift']  #Not used but tidier

        CMFinfo['Mshift'] = float(self.mainCMFgen.Mshift.text())
        self.M_shift = CMFinfo['Mshift']  #Not used but tidier
       
        CMFinfo['mac'] = float(self.mainCMFgen.mac.text())  #Macular and lens densities
        CMFinfo['lens'] = float(self.mainCMFgen.lens.text())

        CMFinfo['linAbs'] = self.mainCMFgen.linAbs.isChecked() #Flags for output file types
        CMFinfo['logAbs'] = self.mainCMFgen.logAbs.isChecked()
        CMFinfo['linQretina'] = self.mainCMFgen.linQretina.isChecked()
        CMFinfo['logQretina'] = self.mainCMFgen.logQretina.isChecked()
        CMFinfo['linQcornea'] = self.mainCMFgen.linQcornea.isChecked()
        CMFinfo['logQcornea'] = self.mainCMFgen.logQcornea.isChecked()
        CMFinfo['linEcornea'] = self.mainCMFgen.linEcornea.isChecked()  
        CMFinfo['logEcornea'] = self.mainCMFgen.logEcornea.isChecked()
        
        CMFinfo['RGBCMFs'] = self.mainCMFgen.RGBCMFs.isChecked()
        
        CMFinfo['Rnm'] = float(self.mainCMFgen.Rnm.text())   #R primary
        CMFinfo['Gnm'] = float(self.mainCMFgen.Gnm.text())   #G primary
        CMFinfo['Bnm'] = float(self.mainCMFgen.Bnm.text())   #B primary

        CMFinfo['DirName'] = str(self.mainCMFgen.DirName.text()) #Directory name

        CMFinfo['FileName'] = str(self.mainCMFgen.FileName.text()) #File and directory name

        CMFinfo['OverwriteYN'] = self.mainCMFgen.OverwriteYN.isChecked() #Overwrite?

        return CMFinfo


    def slotuseCodonsClicked(self):  #Show codons screen and offer button for direct input
        self.LM_codons.show()
        self.useCodons = True   #Set flag
        self.mainCMFgen.useCodons.hide()
        self.mainCMFgen.notuseCodons.show()
        self.mainCMFgen.Lshift.setReadOnly(True)  #Disallow direct text input
        self.mainCMFgen.Mshift.setReadOnly(True)  #Disallow direct text input
        return

 
    def slotresetClicked(self):  #Show codons screen and offer button for directinput
        # Reset radio buttons
        self.LM_codons.M_116_Tyr.setChecked(True)
        self.LM_codons.M_180_Ala.setChecked(True)
        self.LM_codons.M_230_Thr.setChecked(True)
        self.LM_codons.M_233_Ser.setChecked(True)
        self.LM_codons.M_277_Phe.setChecked(True)
        self.LM_codons.M_285_Ala.setChecked(True)
        self.LM_codons.M_309_Phe.setChecked(True)
        self.LM_codons.L_116_Ser.setChecked(True)
        self.LM_codons.L_180_Ser.setChecked(True)
        self.LM_codons.L_230_Ile.setChecked(True)
        self.LM_codons.L_233_Ala.setChecked(True)
        self.LM_codons.L_277_Tyr.setChecked(True)
        self.LM_codons.L_285_Thr.setChecked(True)
        self.LM_codons.L_309_Tyr.setChecked(True)
            
        self.LM_codons.hide()   #Hide codons screen
        self.useCodons = False   #Set flag for not use codons
        self.mainCMFgen.useCodons.show()
        self.mainCMFgen.notuseCodons.hide()
  
        self.M_shift = 0  #Reseet all shifts
        self.L_shift = 0
        self.ML_shift = 0  #For codon screen calculations
        self.LM_shift = 0  #For codon screen calculations


        #Blank some messages and warnings and buttons for later use
        self.mainCMFgen.notuseCodons.hide() #Button to revert to initial default of using direct input for L and M shifts
        self.mainCMFgen.useMforL.setVisible(False)  #Notification that M-cone template is being used for L (hidden)
        self.mainCMFgen.useLforM.setVisible(False)  #Notification that L-cone template is being used for M (hidden)
  
        self.mainCMFgen.useMforL.setVisible(False)  #Notification that M-cone template is being used for L (hidden)
        self.mainCMFgen.useLforM.setVisible(False)  #Notification that L-cone template is being used for M (hidden)
        
        self.mainCMFgen.Lod.setText(str(self.Lod_norm_2))  #Reset input screen to defaults values
        self.mainCMFgen.Mod.setText(str(self.Mod_norm_2))
        self.mainCMFgen.Sod.setText(str(self.Sod_norm_2))
        self.mainCMFgen.mac.setText(str(self.mac_norm_2_460))
        self.mainCMFgen.lens.setText(str(self.lens_norm_400))
         #Zero shifts loaded into window. S cone spectral sensitivity assumed to be fixed.
        self.mainCMFgen.Lshift.setText(str(0.0))
        self.mainCMFgen.Mshift.setText(str(0.0))
        
        return


    def slotnotuseCodonsClicked(self):  #Hide codon screen and offer button to use codons again
        self.LM_codons.hide()
        self.useCodons = False  #Unset flag
        self.mainCMFgen.notuseCodons.hide()
        self.mainCMFgen.useCodons.show()
        self.mainCMFgen.Lshift.setReadOnly(False)   #Allow direct text input
        self.mainCMFgen.Mshift.setReadOnly(False)   #Allow direct text input
        
        return


    def deg2DensitiesClicked(self):
    
        self.mainCMFgen.Lod.setText(str(self.Lod_norm_2))
        self.mainCMFgen.Mod.setText(str(self.Mod_norm_2))
        self.mainCMFgen.Sod.setText(str(self.Sod_norm_2))
        self.mainCMFgen.mac.setText(str(self.mac_norm_2_460))
        self.mainCMFgen.lens.setText(str(self.lens_norm_400))
        
        return
   
    
    def deg10DensitiesClicked(self):
        self.mainCMFgen.Lod.setText(str(self.Lod_norm_10))
        self.mainCMFgen.Mod.setText(str(self.Mod_norm_10))
        self.mainCMFgen.Sod.setText(str(self.Sod_norm_10))
        self.mainCMFgen.mac.setText(str(self.mac_norm_10_460))
        self.mainCMFgen.lens.setText(str(self.lens_norm_400))
        
        return
        
 
    def slotcodonsDoneClicked(self):
        #Calculation done separately from Codon window real-time calculations but checked against them
        #Calculate Mshift for codons changing from M to L (otherwise 0)
        self.M_shift = 0
        if self.LM_codons.M_116_Ser.isChecked():
            self.M_shift = self.M_shift +  self.M_L_116shift
        if self.LM_codons.M_180_Ser.isChecked():
            self.M_shift = self.M_shift + self.M_L_180shift 
        if self.LM_codons.M_230_Ile.isChecked():
            self.M_shift = self.M_shift + self.M_L_230shift
        if self.LM_codons.M_233_Ala.isChecked():
            self.M_shift = self.M_shift + self.M_L_233shift
        if self.LM_codons.M_277_Tyr.isChecked():
            self.M_shift = self.M_shift + self.M_L_277shift
        if self.LM_codons.M_285_Thr.isChecked(): 
            self.M_shift = self.M_shift + self.M_L_285shift
        if self.LM_codons.M_309_Tyr.isChecked():
            self.M_shift = self.M_shift + self.M_L_309shift
      
        self.temp=round(self.M_shift, 3)                  #Round for text
        self.mainCMFgen.Mshift.setText(str(self.temp))    #Write in window
        
        print (self.M_shift)
        print (str(self.M_shift))

        if abs(self.M_shift - self.ML_shift) < 0.001:   #Check this calculation against realtime window calculation == too precise
            print ("M shift calculations tally.")
        else:
            print ("M shift calculations do not tally!")
            print (self.M_shift)
            print (self.ML_shift)

        #Calculate Lshift for codons changing from L to M (otherwise 0)
        self.L_shift = 0
        if self.LM_codons.L_116_Tyr.isChecked():
            self.L_shift = self.L_shift + self.L_M_116shift
        if self.LM_codons.L_180_Ala.isChecked():
            self.L_shift = self.L_shift + self.L_M_180shift 
        if self.LM_codons.L_230_Thr.isChecked():
            self.L_shift = self.L_shift + self.L_M_230shift
        if self.LM_codons.L_233_Ser.isChecked():
            self.L_shift = self.L_shift + self.L_M_233shift
        if self.LM_codons.L_277_Phe.isChecked():
            self.L_shift = self.L_shift + self.L_M_277shift       
        if self.LM_codons.L_285_Ala.isChecked():
            self.L_shift = self.L_shift + self.L_M_285shift
        if self.LM_codons.L_309_Phe.isChecked():
            self.L_shift = self.L_shift + self.L_M_309shift
        self.temp=round(self.L_shift, 3)                  #Round for text
        self.mainCMFgen.Lshift.setText(str(self.temp))    #Write in window

        print (self.L_shift)
        print (str(self.L_shift))


        if abs(self.L_shift - self.LM_shift) < 0.001:   #Check this calculation against realtime window calculation == too precise
            print ("L shift calculations tally.")
        else:
            print ("L shift calculations do not tally!")
            print (self.L_shift)
            print (self.LM_shift)
        self.LM_codons.hide()
        
        self.chooseLMtemplates()
        
        return

    #Codon radio button event handlers
    # M
    def radioM_116_Ser_changed(self):
        if self.LM_codons.M_116_Ser.isChecked():
            self.ML_shift = self.ML_shift + self.M_L_116shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        if self.LM_codons.M_116_Tyr.isChecked():
            self.ML_shift = self.ML_shift - self.M_L_116shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        return

    def radioM_180_Ser_changed(self):
        if self.LM_codons.M_180_Ser.isChecked():
            self.ML_shift = self.ML_shift + self.M_L_180shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        if self.LM_codons.M_180_Ala.isChecked():
            self.ML_shift = self.ML_shift - self.M_L_180shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        return

    def radioM_230_Ile_changed(self):
        if self.LM_codons.M_230_Ile.isChecked():
            self.ML_shift = self.ML_shift + self.M_L_230shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        if self.LM_codons.M_230_Thr.isChecked():
            self.ML_shift = self.ML_shift - self.M_L_230shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        return

    def radioM_233_Ala_changed(self):
        if self.LM_codons.M_233_Ala.isChecked():
            self.ML_shift = self.ML_shift + self.M_L_233shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        if self.LM_codons.M_233_Ser.isChecked():
            self.ML_shift = self.ML_shift - self.M_L_233shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        return

    def radioM_277_Tyr_changed(self):
        if self.LM_codons.M_277_Tyr.isChecked():
            self.ML_shift = self.ML_shift + self.M_L_277shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        if self.LM_codons.M_277_Phe.isChecked():
            self.ML_shift = self.ML_shift - self.M_L_277shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        return

    def radioM_285_Thr_changed(self):
        if self.LM_codons.M_285_Thr.isChecked():
            self.ML_shift = self.ML_shift + self.M_L_285shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        if self.LM_codons.M_285_Ala.isChecked():
            self.ML_shift = self.ML_shift - self.M_L_285shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        return

    def radioM_309_Tyr_changed(self):
        if self.LM_codons.M_309_Tyr.isChecked():
            self.ML_shift = self.ML_shift + self.M_L_309shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        if self.LM_codons.M_309_Phe.isChecked():
            self.ML_shift = self.ML_shift - self.M_L_309shift
            self.temp=round(self.ML_shift, 3)
            self.LM_codons.ML_shift.setText(str(self.temp))
        return
        
    # L        
    def radioL_116_Tyr_changed(self):
        if self.LM_codons.L_116_Tyr.isChecked():
            self.LM_shift = self.LM_shift + self.L_M_116shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        if self.LM_codons.L_116_Ser.isChecked():
            self.LM_shift = self.LM_shift - self.L_M_116shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        return

    def radioL_180_Ala_changed(self):
        if self.LM_codons.L_180_Ala.isChecked():
            self.LM_shift = self.LM_shift + self.L_M_180shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        if self.LM_codons.L_180_Ser.isChecked():
            self.LM_shift = self.LM_shift - self.L_M_180shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        return

    def radioL_230_Thr_changed(self):
        if self.LM_codons.L_230_Thr.isChecked():
            self.LM_shift = self.LM_shift + self.L_M_230shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        if self.LM_codons.L_230_Ile.isChecked():
            self.LM_shift = self.LM_shift - self.L_M_230shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        return

    def radioL_233_Ser_changed(self):
        if self.LM_codons.L_233_Ser.isChecked():
            self.LM_shift = self.LM_shift + self.L_M_233shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        if self.LM_codons.L_233_Ala.isChecked():
            self.LM_shift = self.LM_shift - self.L_M_233shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        return

    def radioL_277_Phe_changed(self):
        if self.LM_codons.L_277_Phe.isChecked():
            self.LM_shift = self.LM_shift + self.L_M_277shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        if self.LM_codons.L_277_Tyr.isChecked():
            self.LM_shift = self.LM_shift - self.L_M_277shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        return

    def radioL_285_Ala_changed(self):
        if self.LM_codons.L_285_Ala.isChecked():
            self.LM_shift = self.LM_shift + self.L_M_285shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        if self.LM_codons.L_285_Thr.isChecked():
            self.LM_shift = self.LM_shift - self.L_M_285shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        return

    def radioL_309_Phe_changed(self):
        if self.LM_codons.L_309_Phe.isChecked():
            self.LM_shift = self.LM_shift + self.L_M_309shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        if self.LM_codons.L_309_Tyr.isChecked():
            self.LM_shift = self.LM_shift - self.L_M_309shift
            self.temp=round(self.LM_shift, 3)
            self.LM_codons.LM_shift.setText(str(self.temp))
        return


    def chooseLMtemplates(self):
        
        if self.useCodons == False: #using lmax inputs
            if self.mainCMFgen.CMFinfo['Mshift'] >= (self.M_L_277shift + self.M_L_285shift):   # Exon 5 = L
                self.MisL = True          # Set use L template for M 
                self.MisLshift = self.mainCMFgen.CMFinfo['Mshift']-(self.Lserlmax_template-self.Mlmax_template)  #Set lmax shift for "M" relative to L lmax 
                self.mainCMFgen.useLforM.setVisible(True)
            else:
                self.MisL = False
                self.mainCMFgen.useLforM.setVisible(False) 

            if self.mainCMFgen.CMFinfo['Lshift'] <= (self.L_M_277shift + self.L_M_285shift):  #Exon 5 = M
                self.LisM = True
                self.LisMshift = self.mainCMFgen.CMFinfo['Lshift'] + (self.Lserlmax_template-self.Mlmax_template)    # Set lmax shift for "L" relative to M lmax
                self.mainCMFgen.useMforL.setVisible(True) 
            else:
                self.LisM = False
                self.mainCMFgen.useMforL.setVisible(False) 

        if self.useCodons == True: #Using codon inputs
             if (self.LM_codons.M_277_Tyr.isChecked() and self.LM_codons.M_285_Thr.isChecked()): #Exon 5 = L Set use L template for M
                self.MisL = True          # Set use L template for M 
                self.MisLshift = self.mainCMFgen.CMFinfo['Mshift']-(self.Lserlmax_template-self.Mlmax_template)  #Set lmax shift for "M" relative to L lmax 
                self.mainCMFgen.useLforM.setVisible(True)
             else:
                self.MisL = False
                self.mainCMFgen.useLforM.setVisible(False) 

             if (self.LM_codons.L_277_Phe.isChecked() and self.LM_codons.L_285_Ala.isChecked()): #Exon 5 = M Set use M template for L
                self.LisM = True          # Set use L template for M 
                self.LisMshift = self.mainCMFgen.CMFinfo['Lshift']+(self.Lserlmax_template-self.Mlmax_template)  #Set lmax shift for "L" relative to M lmax 
                self.mainCMFgen.useMforL.setVisible(True)
             else:
                self.LisM = False
                self.mainCMFgen.useMforL.setVisible(False)
              
        return
            
                
    def checkShiftsizes(self):
        
        self.mainCMFgen.CMFinfo = self.readFromGui()
        #CMFinfo range checks and warnings (currently just Lshift and Mshift)       
        #Range of M and L shifts after which there are template values between 390 and 830 nm
        if (self.mainCMFgen.CMFinfo['Mshift'] > 30) or (self.mainCMFgen.CMFinfo['Mshift'] < -20) :
            self.mainCMFgen.MshiftWarning.setVisible(True)
            return
        else:
            self.mainCMFgen.MshiftWarning.setVisible(False)           #Blank it
        #Range of Lshifts after which there are L or M template values between 390 and 830 nm
        if (self.mainCMFgen.CMFinfo['Lshift'] > 10) or (self.mainCMFgen.CMFinfo['Lshift'] < -40):
            self.mainCMFgen.LshiftWarning.setVisible(True)
            return
        else:
            self.mainCMFgen.LshiftWarning.setVisible(False)           #Blank it
        
        return
        

    def calculateSpectralsensitivities(self):  #Calculates LMS using Lser, M and S with shifts. 
        
        #Set up nanometer values for given step size and set up matrix array size for nm and LMS        
        self.nm_step = self.mainCMFgen.CMFinfo['step_size']
        self.nm = np.arange(360.0, 850+self.nm_step, self.nm_step)  #array with nm steps
        self.LMS_array_size = (len(self.nm), 4) #nm,L,M,S array size
   
        self.mac=CMFtemplates.macular(self.nm)  #Generate macular from template
        self.lens=CMFtemplates.lens(self.nm)    #Generate lens from template


        if self.LisM and not self.MisL:      # 2 M-cone templates
            self.coneabs_template = CMFtemplates.MMSconelog(self.nm, self.LMS_array_size, self.LisMshift, self.mainCMFgen.CMFinfo['Mshift'], 0, 'lin')
            self.log_coneabs_template = CMFtemplates.MMSconelog(self.nm, self.LMS_array_size, self.LisMshift, self.mainCMFgen.CMFinfo['Mshift'], 0, 'log')

        elif self.MisL and not self.LisM:
            self.coneabs_template = CMFtemplates.LLSconelog(self.nm, self.LMS_array_size, self.mainCMFgen.CMFinfo['Lshift'], self.MisLshift, 0, 'lin')
            self.log_coneabs_template = CMFtemplates.LLSconelog(self.nm, self.LMS_array_size, self.mainCMFgen.CMFinfo['Lshift'], self.MisLshift, 0, 'log')

        elif self.MisL and self.LisM:
            self.coneabs_template = CMFtemplates.MLSconelog(self.nm, self.LMS_array_size, self.LisMshift, self.MisLshift, 0, 'lin')
            self.log_coneabs_template = CMFtemplates.MLSconelog(self.nm, self.LMS_array_size, self.LisMshift, self.MisLshift, 0, 'log')

        else:
            self.coneabs_template = CMFtemplates.LMSconelog(self.nm, self.LMS_array_size, self.mainCMFgen.CMFinfo['Lshift'], self.mainCMFgen.CMFinfo['Mshift'], 0, 'lin')
            self.log_coneabs_template = CMFtemplates.LMSconelog(self.nm, self.LMS_array_size, self.mainCMFgen.CMFinfo['Lshift'], self.mainCMFgen.CMFinfo['Mshift'], 0, 'log')

   
        #Retinal absorptances
        self.conenewq_retina = CMFcalc.absorptancefromabsorbance(self.coneabs_template, self.mainCMFgen.CMFinfo['Lod'], self.mainCMFgen.CMFinfo['Mod'], self.mainCMFgen.CMFinfo['Sod'], 'lin')
        self.log_conenewq_retina = CMFcalc.absorptancefromabsorbance(self.coneabs_template, self.mainCMFgen.CMFinfo['Lod'], self.mainCMFgen.CMFinfo['Mod'], self.mainCMFgen.CMFinfo['Sod'], 'log')
        #Corneal quantal spectral sensitivities
        self.conenewq_cornea = CMFcalc.corneafromlinabsorptance(self.conenewq_retina, self.mac, self.lens, self.mainCMFgen.CMFinfo['mac'], self.mainCMFgen.CMFinfo['lens'], 'lin')
        self.log_conenewq_cornea = CMFcalc.corneafromlinabsorptance(self.conenewq_retina, self.mac, self.lens, self.mainCMFgen.CMFinfo['mac'], self.mainCMFgen.CMFinfo['lens'], 'log')
        #Corneal energy spectral sensitivities
        self.conenewe_cornea = CMFcalc.energyfromquantalin(self.conenewq_cornea, 'lin')
        self.log_conenewe_cornea = CMFcalc.energyfromquantalog(self.log_conenewq_cornea, 'log')
        
        #RGB CMFs
        self.RGBCMFs=self.calcRGBCMFs(False)    #Calculate RGB CMFs from LMS given three primaries using Lser, M and S.

        return
    
    
    def calculateSpectralsensitivities_Normal(self): #Calculates shifted Lser, M and S functions
        
        #Set up nanometer values for given step size and set up matrix array size for nm and LMS        
        self.nm_step = self.mainCMFgen.CMFinfo['step_size']
        self.nm = np.arange(360.0, 850+self.nm_step, self.nm_step)  #array with nm steps
        self.LMS_array_size = (len(self.nm), 4) #nm,L,M,S array size
   
        self.mac=CMFtemplates.macular(self.nm)  #Generate macular from template
        self.lens=CMFtemplates.lens(self.nm)    #Generate lens from template

        self.coneabs_template_N = CMFtemplates.LMSconelognormal(self.nm, self.LMS_array_size, 'lin')
        self.log_coneabs_template_N = CMFtemplates.LMSconelognormal(self.nm, self.LMS_array_size, 'log')


        #Retinal absorptances
        self.conenewq_retina_N = CMFcalc.absorptancefromabsorbance(self.coneabs_template_N, self.mainCMFgen.CMFinfo['Lod'], self.mainCMFgen.CMFinfo['Mod'], self.mainCMFgen.CMFinfo['Sod'], 'lin')
        self.log_conenewq_retina_N = CMFcalc.absorptancefromabsorbance(self.coneabs_template_N, self.mainCMFgen.CMFinfo['Lod'], self.mainCMFgen.CMFinfo['Mod'], self.mainCMFgen.CMFinfo['Sod'], 'log')
        #Corneal quantal spectral sensitivities
        self.conenewq_cornea_N = CMFcalc.corneafromlinabsorptance(self.conenewq_retina_N, self.mac, self.lens, self.mainCMFgen.CMFinfo['mac'], self.mainCMFgen.CMFinfo['lens'], 'lin')
        self.log_conenewq_cornea_N = CMFcalc.corneafromlinabsorptance(self.conenewq_retina_N, self.mac, self.lens, self.mainCMFgen.CMFinfo['mac'], self.mainCMFgen.CMFinfo['lens'], 'log')
        #Corneal energy spectral sensitivities
        self.conenewe_cornea_N = CMFcalc.energyfromquantalin(self.conenewq_cornea_N, 'lin')
        self.log_conenewe_cornea_N = CMFcalc.energyfromquantalog(self.log_conenewq_cornea_N, 'log')

        #RGB CMFs
        self.RGBCMFs_N=self.calcRGBCMFs(True)   #Calculate unshifted Lmean, M and S functions
        
        return


    def calcRGBCMFs(self, Normal): #Calculate RGB CMFs from LMS given three primaries. Local variables mainly used except for the output. Normal True uses Lmean, M and S unshifted. 
    
        CMF_Rnm = self.mainCMFgen.CMFinfo['Rnm']
        CMF_Gnm = self.mainCMFgen.CMFinfo['Gnm'] 
        CMF_Bnm = self.mainCMFgen.CMFinfo['Bnm']
        
        nmRGB = np.array([CMF_Rnm, CMF_Gnm, CMF_Bnm])

        
        #Set up nanometer values for given step size and set up matrix array size for nm and LMS plus RGB primaries        
        nm_step = self.mainCMFgen.CMFinfo['step_size']
        nm = np.arange(360.0, 850+nm_step, nm_step)  #array with nm steps
        nmplusRGB = np.concatenate((nmRGB, nm)) # Also calculate RGB primary wavevengths to get the correct normalisation
        
        LMSplus_array_size = (len(nmplusRGB), 4) #nm,L,M,S array size for LMS/RGB calculations
              
        macplus=CMFtemplates.macular(nmplusRGB)  #Generate macular from template
        lensplus=CMFtemplates.lens(nmplusRGB)    #Generate lens from template
         
        if Normal==False:
            coneabs_templateplus = CMFtemplates.LMSconelog(nmplusRGB, LMSplus_array_size, self.mainCMFgen.CMFinfo['Lshift'], self.mainCMFgen.CMFinfo['Mshift'], 0, 'lin')
        
        else:
            coneabs_templateplus = CMFtemplates.LMSconelognormal(nmplusRGB, LMSplus_array_size, 'lin')
            print("True")

       
        conenewq_retinaplus = CMFcalc.absorptancefromabsorbance(coneabs_templateplus, self.mainCMFgen.CMFinfo['Lod'], self.mainCMFgen.CMFinfo['Mod'], self.mainCMFgen.CMFinfo['Sod'], 'lin')

        conenewq_corneaplus = CMFcalc.corneafromlinabsorptance(conenewq_retinaplus, macplus, lensplus, self.mainCMFgen.CMFinfo['mac'], self.mainCMFgen.CMFinfo['lens'], 'lin')
        conenewe_corneaplus = CMFcalc.energyfromquantalin(conenewq_corneaplus, 'lin')
        
      
        RGBLMS = np.transpose(conenewe_corneaplus[0:3, 1:4]) #Col 1 LMS R sensitivities; Col 2 LMS G sensitivities; Col 3 LMS B sensitivities 
 
        LMSRGB=np.linalg.inv(RGBLMS)  # LMS to RGB 3x3 transformation
        
        conenewe_corneaplus = conenewe_corneaplus [3:, :] #Strip off first 3 rows containing primary wavelengths

        #Linear transformation column by column...
        RGBCMFs=np.copy(conenewe_corneaplus)  # Correct array size for RGBCMFs with nm in column 1 for output (DO NOT USE = or calcuations affect both!)
        #col 0 nm
        #col 1 R CMF
        RGBCMFs[:,1] = LMSRGB[0,0] * conenewe_corneaplus[:,1] + LMSRGB[0,1] * conenewe_corneaplus[:,2] + LMSRGB[0,2] * conenewe_corneaplus[:,3] 
        #col 2 G CMF
        RGBCMFs[:,2] = LMSRGB[1,0] * conenewe_corneaplus[:,1] + LMSRGB[1,1] * conenewe_corneaplus[:,2] + LMSRGB[1,2] * conenewe_corneaplus[:,3]   
        #col 3 B CMF
        RGBCMFs[:,3] = LMSRGB[2,0] * conenewe_corneaplus[:,1] + LMSRGB[2,1] * conenewe_corneaplus[:,2] + LMSRGB[2,2] * conenewe_corneaplus[:,3]   
       
        return RGBCMFs



    def writeSensitivities(self):    

        if os.path.isdir(self.mainCMFgen.CMFinfo['DirName'])  == False: #Write
            os.mkdir(self.mainCMFgen.CMFinfo['DirName'])  #Make directory for data output       
        
        if os.path.isfile(self.mainCMFgen.CMFinfo['DirName'] + '/' + self.mainCMFgen.CMFinfo['FileName'] + '.xlsx') and self.mainCMFgen.CMFinfo['OverwriteYN'] == False: #Warning
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("Overwrite error")
            msg.setInformativeText('Choose a different filename \nor select Overwrite file')
            msg.setWindowTitle("Warning")
            msg.exec_()
            return

        XLbook=openpyxl.Workbook()
        XLbook.remove(XLbook.worksheets[0])
        
        #Write Excel worksheets according to flags set on input GUI in data output directory
        if self.mainCMFgen.CMFinfo['linAbs']:
            wslinAbs = XLbook.create_sheet("Lin absorbance")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.coneabs_template.tolist():
                wslinAbs.append(row)
            
        if self.mainCMFgen.CMFinfo['logAbs']:
            wslogAbs = XLbook.create_sheet("Log absorbance")
            wslogAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.log_coneabs_template.tolist():
                wslogAbs.append(row)
            
        if self.mainCMFgen.CMFinfo['linQretina']:
            wslinQretina = XLbook.create_sheet("Lin retinal (Quanta)")
            wslinQretina.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.conenewq_retina.tolist():
                wslinQretina.append(row)
            
        if self.mainCMFgen.CMFinfo['logQretina']:
            wslogQretina = XLbook.create_sheet("Log retinal (Quanta)")
            wslogQretina.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.log_conenewq_retina.tolist():
                wslogQretina.append(row)
            
        if self.mainCMFgen.CMFinfo['linQcornea']:
            wslinQcornea = XLbook.create_sheet("Lin corneal (Quanta)")
            wslinQcornea.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.conenewq_cornea.tolist():
                wslinQcornea.append(row)
            
        if self.mainCMFgen.CMFinfo['logQcornea']:
            wslogQcornea = XLbook.create_sheet("Log corneal (Quanta)")
            wslogQcornea.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.log_conenewq_cornea.tolist():
                wslogQcornea.append(row)
            
        if self.mainCMFgen.CMFinfo['linEcornea']:
            wslinEcornea = XLbook.create_sheet("Lin corneal (Energy)")
            wslinEcornea.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.conenewe_cornea.tolist():
                wslinEcornea.append(row)
            
        if self.mainCMFgen.CMFinfo['logEcornea']:
            wslogEcornea = XLbook.create_sheet("Log corneal (Energy)")
            wslogEcornea.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.log_conenewe_cornea.tolist():
                wslogEcornea.append(row)
            
        if self.mainCMFgen.CMFinfo['RGBCMFs']:
            wsRGBCMFs = XLbook.create_sheet("RGB CMFs")
            wsRGBCMFs.append(['Wavelength', 'R', 'G', 'B'])
            for row in self.RGBCMFs.tolist():
                wsRGBCMFs.append(row)
                

        if len(XLbook.sheetnames)>0:
            XLbook.save(self.mainCMFgen.CMFinfo['DirName'] + '/' + self.mainCMFgen.CMFinfo['FileName'] + '.xlsx')
            print('Excel file saved')
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("No functions selected")
            msg.setWindowTitle("Warning")
            msg.exec_()
            print('No functions selected')
        
        return
    
    

    def writeSensitivities_Normal(self):
 
        if os.path.isdir(self.mainCMFgen.CMFinfo['DirName'])  == False: #Write
            os.mkdir(self.mainCMFgen.CMFinfo['DirName'])  #Make directory for data output       
            
        if os.path.isfile(self.mainCMFgen.CMFinfo['DirName'] + '/' + self.mainCMFgen.CMFinfo['FileName'] + '.xlsx') and self.mainCMFgen.CMFinfo['OverwriteYN'] == False: #Warning
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("Overwrite error")
            msg.setInformativeText('Choose a different filename \nor select Overwrite file')
            msg.setWindowTitle("Warning")
            msg.exec_()
            return
        
        XLbook=openpyxl.Workbook()
        XLbook.remove(XLbook.worksheets[0])

        #Write Excel worksheets according to flags set on input GUI in data output directory
        if self.mainCMFgen.CMFinfo['linAbs']:
            wslinAbs = XLbook.create_sheet("Lin absorbance")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.coneabs_template_N.tolist():
                wslinAbs.append(row)

        if self.mainCMFgen.CMFinfo['logAbs']:
            wslinAbs = XLbook.create_sheet("Log absorbance")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.log_coneabs_template_N.tolist():
                wslinAbs.append(row)
                
        if self.mainCMFgen.CMFinfo['linQretina']:
            wslinAbs = XLbook.create_sheet("Lin retinal (Quanta)")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.conenewq_retina_N.tolist():
                wslinAbs.append(row)
                
        if self.mainCMFgen.CMFinfo['logQretina']:
            wslinAbs = XLbook.create_sheet("Log retinal (Quanta)")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.log_conenewq_retina_N.tolist():
                wslinAbs.append(row)
                
        if self.mainCMFgen.CMFinfo['linQcornea']:
            wslinAbs = XLbook.create_sheet("Lin corneal (Quanta)")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.conenewq_cornea_N.tolist():
                wslinAbs.append(row)

        if self.mainCMFgen.CMFinfo['logQcornea']:
            wslinAbs = XLbook.create_sheet("Log corneal (Quanta)")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.log_conenewq_cornea_N.tolist():
                wslinAbs.append(row)

        if self.mainCMFgen.CMFinfo['linEcornea']:
            wslinAbs = XLbook.create_sheet("Lin corneal (Energy)")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.conenewe_cornea_N.tolist():
                wslinAbs.append(row)

        if self.mainCMFgen.CMFinfo['logEcornea']:
            wslinAbs = XLbook.create_sheet("Log corneal (Energy)")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.log_conenewe_cornea_N.tolist():
                wslinAbs.append(row)

        if self.mainCMFgen.CMFinfo['RGBCMFs']:
            wslinAbs = XLbook.create_sheet("RGB CMFs")
            wslinAbs.append(['Wavelength', 'L', 'M', 'S'])
            for row in self.RGBCMFs_N.tolist():
                wslinAbs.append(row)
            
        if len(XLbook.sheetnames)>0:
            XLbook.save(self.mainCMFgen.CMFinfo['DirName'] + '/' + self.mainCMFgen.CMFinfo['FileName'] + '.xlsx')
            print('Excel file saved')
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("No functions selected")
            msg.setWindowTitle("Warning")
            msg.exec_()
            print('No functions selected')
        
        return
    

    def slotOutputCMFsClicked(self): #Calculate and write cone sensitivities using Lser, M and S

        self.checkShiftsizes()
        
        self.chooseLMtemplates()   #Choose templates L or M

        self.calculateSpectralsensitivities()

        self.writeSensitivities()
        
        return


    def slotOutputMeanLMSCMFsClicked(self): #Calculate and write Lmean, M andnensitivities

       self.mainCMFgen.CMFinfo = self.readFromGui()
       
       self.calculateSpectralsensitivities_Normal()

       self.writeSensitivities_Normal()

       return
   
       
    def slotPlotCMFsClicked(self):           #Plot new LMS CMFs and compare with normal LMS CMFs
 
        self.mainCMFgen.CMFinfo = self.readFromGui()

        self.checkShiftsizes()

        self.chooseLMtemplates()

        self.calculateSpectralsensitivities()  #Lser, M and S with shifts. Also calculates RGB CMFs from LMS given three primaries using Lser, M and S.

        self.calculateSpectralsensitivities_Normal()  #Lmean, M and S unshifted. Also calculates RGB CMFs from LMS given three primaries using Lser, M and S.

        #Number of plots set to 6: 3 x 2
      
        plotCMF = CMFPlot(rows = 3, cols = 3, nmstepsize=self.nm_step) #Substantiate plotting class (Step size used for chromaticity diagram)

        #Plot new functions and compare with normal. Last parameter is position in subplot array set up in class call (3x3) 1,2,3 top row; 4,5,6 middle row, etx
        plotCMF.LMScompare(self.coneabs_template[:,0], self.coneabs_template[:,1], self.coneabs_template[:,2], self.coneabs_template[:,3], self.coneabs_template_N[:,1], self.coneabs_template_N[:,2], self.coneabs_template_N[:,3], 0, 1.1, 'Absorbance from template', 1)
        plotCMF.LMScompare(self.log_coneabs_template[:,0], self.log_coneabs_template[:,1], self.log_coneabs_template[:,2], self.log_coneabs_template[:,3], self.log_coneabs_template_N[:,1], self.log_coneabs_template_N[:,2], self.log_coneabs_template_N[:,3], -3, 0.2, 'Log absorbance from template', 4)
        plotCMF.LMScompare(self.conenewe_cornea[:,0], self.conenewe_cornea[:,1], self.conenewe_cornea[:,2], self.conenewe_cornea[:,3], self.conenewe_cornea_N[:,1], self.conenewe_cornea_N[:,2], self.conenewe_cornea_N[:,3], 0, 1.1, 'Energy corneal spectral sensitivity', 2)
        plotCMF.LMScompare(self.log_conenewe_cornea[:,0], self.log_conenewe_cornea[:,1], self.log_conenewe_cornea[:,2], self.log_conenewe_cornea[:,3], self.log_conenewe_cornea_N[:,1], self.log_conenewe_cornea_N[:,2], self.log_conenewe_cornea_N[:,3], -3, 0.2, 'Log energy corneal spectral sensitivity', 5)

        #Chromaticities and L-M
        plotCMF.LMchromaticity(self.conenewe_cornea, 'LM chromaticities', 3) #Plot chromaticities
        plotCMF.LoverMcompare(self.conenewe_cornea[:,0], self.conenewe_cornea[:,1], self.conenewe_cornea[:,2], self.conenewe_cornea_N[:,1], self.conenewe_cornea_N[:,2], -0.5, 1, 'L-M', 6) #Plot 2xL/M differences

        plotCMF.RGBCMFsPlot(self.RGBCMFs, 'RGB CMFs', 7) #Plot RGB CMFs 

        #Spyder default is to plot inline. To plot in window Tools->Preferences->IPython console->Graphics->Graphics backend->Automatic
        plotCMF.displayAll(100,100,round(screenwidth*0.6),round(screenheight*0.8), False) #Main window parameters x, y top left corner, width, height, no Hold (same event manager as GUI)

        return

#Run input dialogue--mainly in CSFinput module 

app = QtWidgets.QApplication(sys.argv)
resol=app.desktop().screenGeometry()
screenwidth, screenheight = resol.width(), resol.height()

startWindows = individualtemplatesCMFs()


app.exec_()   #Loop for gui interaction

